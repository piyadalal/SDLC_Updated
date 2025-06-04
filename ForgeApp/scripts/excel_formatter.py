from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def auto_format_excel(path: str):
    wb = load_workbook(path)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 5  # add padding

    wb.save(path)
    print(f"✅ Excel formatted and saved to: {path}")
